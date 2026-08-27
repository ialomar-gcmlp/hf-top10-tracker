#!/usr/bin/env python3
"""
Render the weekly top-10 snapshot: tab-separated blocks for pasting straight
into Excel, plus a .csv and .xlsx for anyone who wants a file.

Input is a run manifest (JSON) assembled by the skill:

{
  "generated": "2026-08-27",
  "funds": [
    {
      "fund": "Coatue Management",
      "source": "WhaleWisdom 13F",          # or "DealCloud ARS Research Note" / "None"
      "as_of": "2026-06-30",
      "detail": "Coatue Management LP (filer 8271), filed 2026-08-14",
      "caveat": "13F long book only ...",   # optional, printed under the header
      "not_found_reason": "...",            # optional, when holdings are short
      "holdings": [{"position": "Meta Platforms", "ticker": "META",
                    "pct_of_portfolio": 9.4}, ...]
    }
  ]
}

Every block is always 10 rows. Missing positions are written as "Not found" --
never padded with filler and never estimated.
"""

import argparse
import csv
import json
import os
import sys

ROWS = 10
NOT_FOUND = "Not found"

GLOBAL_CAVEAT = (
    "13F % of portfolio = % of the manager 13F-REPORTABLE LONG US-LISTED EQUITY book, "
    "NOT % of total AUM. 13Fs exclude shorts, non-US listings, private holdings and cash, "
    "so for managers with large private/short/non-US exposure (D1, Tiger Global, Coatue, "
    "Aspex and similar) these weights overstate the position versus the true total portfolio."
)


def fmt_pct(value):
    return f"{value:.2f}%" if isinstance(value, (int, float)) else NOT_FOUND


def position_label(h):
    name = (h.get("position") or "").strip()
    ticker = (h.get("ticker") or "").strip()
    if not name:
        return NOT_FOUND
    return f"{name} ({ticker})" if ticker else name


def padded_rows(fund):
    """Exactly ROWS (label, pct) pairs, padded with Not found."""
    out = []
    for h in (fund.get("holdings") or [])[:ROWS]:
        out.append((position_label(h), fmt_pct(h.get("pct_of_portfolio"))))
    while len(out) < ROWS:
        out.append((NOT_FOUND, NOT_FOUND))
    return out


def header_line(fund):
    return " | ".join(
        [
            fund.get("fund", "?"),
            fund.get("source") or "No source",
            fund.get("as_of") or "no as-of date",
        ]
    )


def render_tsv(manifest):
    lines = [
        f"HEDGE FUND TOP 10 HOLDINGS -- generated {manifest.get('generated', '')}",
        f"NOTE: {GLOBAL_CAVEAT}",
        "",
    ]
    for fund in manifest["funds"]:
        lines.append(header_line(fund))
        if fund.get("detail"):
            lines.append(f"  source detail: {fund['detail']}")
        if fund.get("caveat"):
            lines.append(f"  caveat: {fund['caveat']}")
        if fund.get("not_found_reason"):
            lines.append(f"  why blank: {fund['not_found_reason']}")
        lines.append("Position\t% of Portfolio")
        for label, pct in padded_rows(fund):
            lines.append(f"{label}\t{pct}")
        lines.append("")
    return "\n".join(lines)


def write_csv(manifest, path):
    """Wide layout: one column-pair per fund, mirroring the paste blocks."""
    funds = manifest["funds"]
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow([f"Hedge fund top 10 holdings - generated {manifest.get('generated','')}"])
        w.writerow([GLOBAL_CAVEAT])
        w.writerow([])

        row_fund, row_meta, row_hdr = [], [], []
        for f in funds:
            row_fund += [f.get("fund", "?"), ""]
            row_meta += [f"{f.get('source') or 'No source'} | {f.get('as_of') or 'no as-of date'}", ""]
            row_hdr += ["Position", "% of Portfolio"]
        w.writerow(row_fund)
        w.writerow(row_meta)
        w.writerow(row_hdr)

        grids = [padded_rows(f) for f in funds]
        for i in range(ROWS):
            row = []
            for g in grids:
                row += [g[i][0], g[i][1]]
            w.writerow(row)

        w.writerow([])
        w.writerow(["Notes"])
        for f in funds:
            bits = [f.get("detail"), f.get("caveat"), f.get("not_found_reason")]
            detail = " / ".join([b for b in bits if b])
            if detail:
                w.writerow([f.get("fund", "?"), detail])
    return path


def _safe_sheet(name, used):
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    name = name[:31] or "Fund"
    base, n = name, 2
    while name in used:
        suffix = f"~{n}"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def write_xlsx(manifest, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("openpyxl not installed -- skipping .xlsx (csv and tsv still produced).",
              file=sys.stderr)
        return None

    funds = manifest["funds"]
    wb = Workbook()
    bold = Font(bold=True)

    # Summary sheet: every fund side by side.
    ws = wb.active
    ws.title = "All Funds"
    ws["A1"] = f"Hedge fund top 10 holdings - generated {manifest.get('generated','')}"
    ws["A1"].font = bold
    ws["A2"] = GLOBAL_CAVEAT
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 46

    for idx, f in enumerate(funds):
        c = 1 + idx * 2
        ws.cell(row=4, column=c, value=f.get("fund", "?")).font = bold
        ws.cell(row=5, column=c,
                value=f"{f.get('source') or 'No source'} | {f.get('as_of') or 'no as-of date'}")
        ws.cell(row=6, column=c, value="Position").font = bold
        ws.cell(row=6, column=c + 1, value="% of Portfolio").font = bold
        for i, h in enumerate(padded_rows(f)):
            ws.cell(row=7 + i, column=c, value=h[0])
            ws.cell(row=7 + i, column=c + 1, value=h[1])
        ws.column_dimensions[ws.cell(row=6, column=c).column_letter].width = 34
        ws.column_dimensions[ws.cell(row=6, column=c + 1).column_letter].width = 14

    # One sheet per fund, with numeric percentages so Excel can do maths on them.
    used = {"All Funds"}
    for f in funds:
        s = wb.create_sheet(_safe_sheet(f.get("fund", "Fund"), used))
        s["A1"] = f.get("fund", "?")
        s["A1"].font = bold
        s["A2"] = f"Source: {f.get('source') or 'No source'}"
        s["A3"] = f"As-of: {f.get('as_of') or 'no as-of date'}"
        if f.get("detail"):
            s["A4"] = f"Detail: {f['detail']}"
        if f.get("caveat"):
            s["A5"] = f"Caveat: {f['caveat']}"
        if f.get("not_found_reason"):
            s["A6"] = f"Why blank: {f['not_found_reason']}"
        s["A8"], s["B8"] = "Position", "% of Portfolio"
        s["A8"].font = bold
        s["B8"].font = bold
        holdings = (f.get("holdings") or [])[:ROWS]
        for i in range(ROWS):
            r = 9 + i
            if i < len(holdings) and isinstance(holdings[i].get("pct_of_portfolio"), (int, float)):
                s.cell(row=r, column=1, value=position_label(holdings[i]))
                cell = s.cell(row=r, column=2, value=holdings[i]["pct_of_portfolio"] / 100.0)
                cell.number_format = "0.00%"
            else:
                s.cell(row=r, column=1,
                       value=position_label(holdings[i]) if i < len(holdings) else NOT_FOUND)
                s.cell(row=r, column=2, value=NOT_FOUND)
        s.column_dimensions["A"].width = 40
        s.column_dimensions["B"].width = 16

    wb.save(path)
    return path


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--manifest", required=True, help="Run manifest JSON, or - for stdin")
    p.add_argument("--outdir", default=".")
    p.add_argument("--basename", default="hf_top10")
    p.add_argument("--tsv-only", action="store_true")
    a = p.parse_args()

    raw = sys.stdin.read() if a.manifest == "-" else open(a.manifest, encoding="utf-8").read()
    manifest = json.loads(raw)

    print(render_tsv(manifest))

    if a.tsv_only:
        return
    os.makedirs(a.outdir, exist_ok=True)
    stamp = manifest.get("generated", "")
    base = os.path.join(a.outdir, f"{a.basename}_{stamp}" if stamp else a.basename)
    written = [write_csv(manifest, base + ".csv"), write_xlsx(manifest, base + ".xlsx")]
    print("\nFiles written:", file=sys.stderr)
    for w in written:
        if w:
            print("  " + os.path.abspath(w), file=sys.stderr)


if __name__ == "__main__":
    main()
